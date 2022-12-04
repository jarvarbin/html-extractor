from bs4 import BeautifulSoup
import requests
import openpyxl

# Set the URL of the website you want to scrape
url = "https://www.example.com"

# Send a GET request to the URL and retrieve the HTML response
response = requests.get(url)

# Use BeautifulSoup to parse the HTML response
soup = BeautifulSoup(response.text, 'html.parser')

# Create an Excel workbook and sheet to save the data
wb = openpyxl.Workbook()
sheet = wb.active

# Set the column headers in the Excel sheet
sheet.cell(row=1, column=1).value = "Link"
sheet.cell(row=1, column=2).value = "HTML Comment"
sheet.cell(row=1, column=3).value = "JS"
sheet.cell(row=1, column=4).value = "URL"

# Initialize the row counter
row_count = 2

# Extract all the links, HTML comments, JS, and URLs from the website and save them to the Excel sheet
for link in soup.find_all('a'):
    sheet.cell(row=row_count, column=1).value = link.get('href')
    row_count += 1

for js in soup.find_all('script'):
    sheet.cell(row=row_count, column=3).value = js.get('src')
    row_count += 1

for url in soup.find_all('link'):
    sheet.cell(row=row_count, column=4).value = url.get('href')
    row_count += 1

# Save the Excel workbook
wb.save("data.xlsx")
